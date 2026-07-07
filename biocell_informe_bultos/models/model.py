# -*- coding: utf-8 -*-
from odoo.tools.misc import DEFAULT_SERVER_DATETIME_FORMAT
import time
import odoo.addons.decimal_precision as dp
from odoo.exceptions import UserError
import base64
from odoo import models, fields, api
import codecs
import subprocess
import sys
from datetime import datetime, timedelta

def install(package):
	subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
	import openpyxl
except:
	install('openpyxl==3.0.5')

import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl import Workbook
values = {}
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell
import datetime
import calendar


class report_package_list(models.TransientModel):
    _name = "report.package.list"
    _description = "report.package.list"

    def get_report(self):

        cad = ""
        s_prod = [-1, -1, -1]
        s_loca = [-1, -1, -1]
        import io
        import base64
        from xlsxwriter.workbook import Workbook
        import datetime

        output = io.BytesIO()
        workbook = Workbook(output, {'constant_memory': False})
        fechashow = str(  datetime.datetime.now() - datetime.timedelta(hours=5)  )

        worksheet = workbook.add_worksheet("reporte de Packs List")


        bold_border = workbook.add_format({'bold': True, 'bg_color': '#D9F1FF'})

        bold_border.set_font_size(11)
        bold_border.set_border(style=1)
        #bold_border.set_text_wrap()  # Ajuste de texto


        bold_border_texto = workbook.add_format({'bold': False})
        bold_border_texto.set_font_size(10)
        bold_border_texto.set_border(style=1)
        #bold_border_texto.set_text_wrap()  # Ajuste de texto


        bold_border_entero = workbook.add_format({'bold': False})
        bold_border_entero.set_font_size(10)
        bold_border_entero.set_border(style=1)
        bold_border_entero.set_num_format('#,##0')  # Formato de número entero



        especial1 = workbook.add_format({'bold': True, 'bg_color': '#D9F1FF'})
        especial1.set_align('center')
        especial1.set_align('vcenter')
        especial1.set_text_wrap()
        especial1.set_font_size(17)
        #especial1.set_border(style=1)



        worksheet.set_column('A:A', 35)
        worksheet.set_column('B:D', 35)
        worksheet.set_column('E:E', 75)
        worksheet.set_column('F:Z', 35)
        #worksheet.merge_range(1, 0, 2, 1, "", especial1)
        worksheet.merge_range(1, 0, 2, 7, "Reporte de Packs List", especial1)
        worksheet.write(3, 0, "Fecha", bold_border)
        worksheet.write(3, 1, str(fechashow)[:19], bold_border_texto)


        #worksheet.merge_range(17, 2, 17, 3, "PARÁMETROS", bold_border)
        worksheet.write(5, 0, "Paquete", bold_border)
        worksheet.write(5, 1, "Estado del Pack", bold_border)
        worksheet.write(5, 2, "Lugar", bold_border)
        worksheet.write(5, 3, "Plantilla", bold_border)
        worksheet.write(5, 4, "Producto", bold_border)
        worksheet.write(5, 5, "Cod. Producto", bold_border)
        worksheet.write(5, 6, "Lote", bold_border)
        worksheet.write(5, 7, "Cantidad", bold_border)

        x = 6
        #final = []
        fechaformato = workbook.add_format({'bold': False, 'num_format': 'dd/mm/yyyy'})
        fechaformato.set_font_size(10)
        fechaformato.set_border(style=1)
        numeroformato = workbook.add_format({'bold': False, 'num_format': '0.00'})
        numeroformato.set_font_size(10)
        numeroformato.set_border(style=1)

        horaformato = workbook.add_format({'bold': False, 'num_format': 'hh:mm:ss'})


        horaformato.set_font_size(10)
        horaformato.set_border(style=1)
        for linea in self.env['stock.quant.package'].search([]):
            for l in linea.quant_ids:
                worksheet.write(x, 0, linea.name or '', bold_border_texto)
                worksheet.write(x, 1, linea.status_validate or '', bold_border_texto)
                worksheet.write(x, 2, linea.location_id.name_get()[0][1] or '', bold_border_texto)
                worksheet.write(x, 3, linea.template_package_list_id.name, bold_border_texto)
                worksheet.write(x, 4, l.product_id.name_get()[0][1] or "", bold_border_texto)
                worksheet.write(x, 5, l.product_id.default_code or "", bold_border_texto)
                worksheet.write(x, 6, l.lot_id.name or '', bold_border_texto)
                worksheet.write(x, 7, l.quantity or 0, numeroformato)
                x+=1
        workbook.close()

        output.seek(0)
        output_datas = base64.b64encode(output.getvalue())
        output.close()

        ahora = datetime.datetime.now()
        return self.env['popup.it'].get_file('Reporte de Packs List' + '.xlsx', output_datas)


class report_package_incomplete(models.TransientModel):
    _name = "report.package.incomplete"
    _description = "report.package.incomplete"

    #fecha_inicio = fields.Date(string="Fecha Inicio", required=True)
    #fecha_fin = fields.Date(string="Fecha Fin", required=True)

    def get_report(self):
        #fecha_inicio = self.fecha_inicio
        #fecha_fin = self.fecha_fin


        cad = ""
        s_prod = [-1, -1, -1]
        s_loca = [-1, -1, -1]
        import io
        import base64
        from xlsxwriter.workbook import Workbook
        import datetime

        output = io.BytesIO()
        workbook = Workbook(output, {'constant_memory': False})
        fechashow = str(  datetime.datetime.now() - datetime.timedelta(hours=5)  )

        worksheet = workbook.add_worksheet("reporte de Packs Incompletos")


        bold_border = workbook.add_format({'bold': True, 'bg_color': '#D9F1FF'})

        bold_border.set_font_size(11)
        bold_border.set_border(style=1)
        #bold_border.set_text_wrap()  # Ajuste de texto


        bold_border_texto = workbook.add_format({'bold': False})
        bold_border_texto.set_font_size(10)
        bold_border_texto.set_border(style=1)
        #bold_border_texto.set_text_wrap()  # Ajuste de texto


        bold_border_entero = workbook.add_format({'bold': False})
        bold_border_entero.set_font_size(10)
        bold_border_entero.set_border(style=1)
        bold_border_entero.set_num_format('#,##0')  # Formato de número entero



        especial1 = workbook.add_format({'bold': True, 'bg_color': '#D9F1FF'})
        especial1.set_align('center')
        especial1.set_align('vcenter')
        especial1.set_text_wrap()
        especial1.set_font_size(17)
        #especial1.set_border(style=1)



        worksheet.set_column('A:A', 35)
        worksheet.set_column('B:D', 35)
        worksheet.set_column('E:E', 75)
        worksheet.set_column('F:Z', 35)
        #worksheet.merge_range(1, 0, 2, 1, "", especial1)
        worksheet.merge_range(1, 0, 2, 7, "Reporte de Packs Incompletos", especial1)
        worksheet.write(3, 0, "Fecha", bold_border)
        worksheet.write(3, 1, str(fechashow)[:19], bold_border_texto)


        #worksheet.merge_range(17, 2, 17, 3, "PARÁMETROS", bold_border)
        worksheet.write(5, 0, "Paquete", bold_border)
        worksheet.write(5, 1, "Tipo de picking", bold_border)
        #worksheet.write(5, 2, "Propietario", bold_border)
        worksheet.write(5, 2, "Lugar", bold_border)
        worksheet.write(5, 3, "Plantilla", bold_border)
        worksheet.write(5, 4, "Producto", bold_border)
        worksheet.write(5, 5, "Cantidad en Plantilla", bold_border)
        worksheet.write(5, 6, "Catidad Faltante", bold_border)

        x = 6
        self.env.cr.execute("""
            SELECT 
                sqp.name AS package_id,
                tipo_picking.name,
                --rp.name,
                sl.complete_name,
                sqpl.name,
                CASE 
                    WHEN pp.default_code IS NOT NULL THEN CONCAT('[', pp.default_code, '] ', pname.new_name)
                    ELSE pname.new_name 
                END AS product_name,

                sqli.cantidad AS cantidad_plantilla,
                --COALESCE(SUM(sq.quantity), 0) AS cantidadquant,
                (sqli.cantidad - COALESCE(SUM(sq.quantity), 0)) AS cantidadfaltante
            FROM stock_quant_package sqp
            left join stock_package_type tipo_picking on tipo_picking.id = sqp.package_type_id
            --left join res_partner rp on rp.id = sqp.owner_id
            left join stock_location sl on sl.id = sqp.location_id


            JOIN stock_quant_package_list sqpl ON sqp.template_package_list_id = sqpl.id
            JOIN stock_quant_package_list_items sqli ON sqli.package_list_id = sqpl.id
            left join product_product pp on pp.id = sqli.product_id
						LEFT JOIN ( SELECT t_pp.id,
                            ((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
                        FROM product_product t_pp
                            JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
                            left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
                            LEFT JOIN product_variant_combination pvc ON pvc.product_product_id = t_pp.id
                            LEFT JOIN product_template_attribute_value ptav ON ptav.id = pvc.product_template_attribute_value_id
                            LEFT JOIN product_attribute_value pav ON pav.id = ptav.product_attribute_value_id
                        GROUP BY t_pp.id) pname ON pname.id = sqli.product_id


            LEFT JOIN stock_quant sq ON sq.package_id = sqp.id AND sq.product_id = sqli.product_id
            WHERE 
                (sqp.company_id = """ + str(self.env.company.id) + """ OR sqp.company_id IS NULL)

            GROUP BY sqp.name, tipo_picking.name, sl.complete_name,
                sqpl.name,
                pp.default_code,
                pname.new_name,
                sqli.cantidad
            HAVING (sqli.cantidad - COALESCE(SUM(sq.quantity), 0)) > 0
            ORDER BY sqp.name, pname.new_name;
        """)
        final = self.env.cr.fetchall()
        #final = []
        fechaformato = workbook.add_format({'bold': False, 'num_format': 'dd/mm/yyyy'})
        fechaformato.set_font_size(10)
        fechaformato.set_border(style=1)
        numeroformato = workbook.add_format({'bold': False, 'num_format': '0.00'})
        numeroformato.set_font_size(10)
        numeroformato.set_border(style=1)

        horaformato = workbook.add_format({'bold': False, 'num_format': 'hh:mm:ss'})


        horaformato.set_font_size(10)
        horaformato.set_border(style=1)
        for linea in final:
            worksheet.write(x, 0, linea[0] if linea[0] else "", bold_border_texto)
            worksheet.write(x, 1, linea[1] if linea[1] else "", bold_border_texto)
            worksheet.write(x, 2, linea[2] if linea[2] else "", bold_border_texto)
            worksheet.write(x, 3, linea[3] if linea[3] else "", bold_border_texto)
            worksheet.write(x, 4, linea[4] if linea[4] else "", bold_border_texto)
            worksheet.write(x, 5, linea[5] if linea[5] else 0, numeroformato)
            worksheet.write(x, 6, linea[6] if linea[6] else 0, numeroformato)
            x+=1
        workbook.close()

        output.seek(0)
        output_datas = base64.b64encode(output.getvalue())
        output.close()

        ahora = datetime.datetime.now()
        return self.env['popup.it'].get_file('Reporte de Packs incompleto' + '.xlsx', output_datas)

