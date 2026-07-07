# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
from io import BytesIO
from datetime import *
import base64
import subprocess
import sys

def install(package):
	subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
	import openpyxl
except:
	install('openpyxl==3.0.5')

import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl import Workbook
values = {}
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell


from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.colors import magenta, red , black , blue, gray, Color, HexColor
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Table
from reportlab.lib.units import  cm,mm
from reportlab.lib.utils import simpleSplit
import decimal

def font_calibri_16_bold(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=12, bold=True)
    return cell
def font_calibri_bold(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=16, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    return cell

def font_calibri_11_bold_upper(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=11, bold=True)
    return cell

def font_calibri_11(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=11, bold=False)
    return cell

def format_cel_calibri_11_bold_centered(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=11, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Configura el relleno de la celda en celeste claro
    fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    cell.fill = fill
    
    # Configura los bordes de la celda
    cell.border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'))
    
    return cell


def format_cel_calibri_11_centered(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=11, bold=False)
    
    # Configura los bordes de la celda
    cell.border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'))
    
    return cell

def format_cel_calibri_11_centered_1(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=11, bold=False)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Configura los bordes de la celda
    cell.border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'))
    
    return cell



def format_cel_calibri_11_number(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=11, bold=False)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = "0.00"  # Establece el formato de número a 2 decimales

    # Configura los bordes de la celda
    cell.border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'))
    
    return cell
class biocell_formato_pedido_venta_report_hj_xlsx(models.AbstractModel):
    _name = 'report.biocell_fhc.biocell_formato_hoja_control_xlsx'
    _inherit = 'report.biocell_generador_excel.abstract'

    def generate_xlsx_report(self, workbook, data, so):

        ws = workbook.create_sheet('Hoja_control')

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 15
        ws.column_dimensions['R'].width = 15
        ws.column_dimensions['S'].width = 15
        ws.column_dimensions['T'].width = 15
        ws.column_dimensions['U'].width = 15
        ws.column_dimensions['V'].width = 15
        ws.column_dimensions['W'].width = 15



        ws.append([""])
        ws.merged_cells.ranges.append('D2:I2')
        tittle = font_calibri_16_bold(ws, "BIOCELLS GENOMICS SOCIEDAD ANONIMA CERRADA - BIOGENOMICS S.A.C.")
        ws.append([(''),(''),(''),tittle])
        ws.merged_cells.ranges.append('D3:I3')
        ruc = font_calibri_16_bold(ws, "RUC: 20607531600")
        ws.append([(''),(''),(''),ruc])
        ws.merged_cells.ranges.append('D4:I4')
        direccion = font_calibri_16_bold(ws, "AV. GUARDIA CIVIL NRO. 1321 DPTO. 1303 URB. VILLA VICTORIA - LIMA SURQUILLO")
        ws.append([(''),(''),(''),direccion])
        ws.merged_cells.ranges.append('D5:I5')
        telefono = font_calibri_16_bold(ws, "Telf: +51 1 6952357")
        ws.append([(''),(''),(''),telefono])
        ws.merged_cells.ranges.append('D6:I6')
        email = font_calibri_16_bold(ws, "Correo Electrónico: biocells.peru@biocells.pe ")
        ws.append([(''),(''),(''),email])
                                  

        ws.append([""])

        ws.merged_cells.ranges.append('B10:C10')
        fact = font_calibri_11_bold_upper(ws,'CÓDIGO DE CIRUGÍA')
        ws.append([(''),fact,(''),font_calibri_11(ws,': '+str(so.sugery_order if so.sugery_order else '')),(''),font_calibri_11_bold_upper(ws,'NOMBRE PACIENTE'),font_calibri_11(ws,': '+ str(so.name_patient))])
        ws.merged_cells.ranges.append('B11:C11')
        calle = font_calibri_11_bold_upper(ws,'FECHA DE CIRGUGÍA')
        ws.append([(''),calle,(''),font_calibri_11(ws,': '+str(so.sugery_date.strftime('%Y-%m-%d') if so.sugery_date else '')),(''),font_calibri_11_bold_upper(ws,'HISTORIA CLÍNICA'),font_calibri_11(ws,': '+ str(so.clinic_history if so.clinic_history else ''))])
        ws.merged_cells.ranges.append('B12:C12')
        atencion = font_calibri_11_bold_upper(ws,'HORA DE CIRGUÍA')
        ws.append([(''),atencion,(''),font_calibri_11(ws,': '+str(so.sugery_date.strftime('%H:%M') if so.sugery_date else '')),(''),font_calibri_11_bold_upper(ws,'DNI/RUC'),font_calibri_11(ws,': '+ str(so.patient_vat if so.patient_vat else ''))])
        ws.merged_cells.ranges.append('B13:C13')
        atencion = font_calibri_11_bold_upper(ws,'CIRUJANO')
        ws.append([(''),atencion,(''),font_calibri_11(ws,': '+str(so.name_doctor.name if so.name_doctor else '')),(''),font_calibri_11_bold_upper(ws,'PROCEDIMIENTO'),font_calibri_11(ws,': '+ str(so.procedure_clinic if so.procedure_clinic else ''))])
        atencion = font_calibri_11_bold_upper(ws,'CENTRO MÉDICO')
        ws.append([(''),atencion,(''),font_calibri_11(ws,': '+str(so.medic_center if so.medic_center else '')),('')])

       
       
        ws.append([""])
        ws.append([(''),format_cel_calibri_11_bold_centered(ws,'CÓDIGO'),format_cel_calibri_11_bold_centered(ws,'DESCRIPCIÓN'),format_cel_calibri_11_bold_centered(ws,'LOTE'),format_cel_calibri_11_bold_centered(ws,'RS'),format_cel_calibri_11_bold_centered(ws,'CANTIDAD'),format_cel_calibri_11_bold_centered(ws,'P. UNIT.'),format_cel_calibri_11_bold_centered(ws,'P. TOTAL.')])
        subtotal=0
        for move in so.move_ids_return:
            for line in move.move_line_ids:
                if line.qty_consumida > 0:
                    ws.append([(''),format_cel_calibri_11_centered_1(ws,str(line.product_id.default_code) if line.product_id.default_code else " "),format_cel_calibri_11_number(ws, str(line.product_id.name)),format_cel_calibri_11_centered_1(ws,str(line.lot_id.name or '')),format_cel_calibri_11_centered_1(ws,str(line.x_studio_registro_sanitario or '')),format_cel_calibri_11_centered_1(ws,str(int(line.qty_consumida))),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(move.price_unit_sale))),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(line.qty_consumida * move.price_unit_sale)))])
                    subtotal = subtotal + (line.qty_consumida * move.price_unit_sale)
        
        ws.append([""])
        ws.append([(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"SUBTOTAL"),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(subtotal)))])
        ws.append([(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"IGV"),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(subtotal*0.18)))])
        ws.append([(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"TOTAL"),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(subtotal+(subtotal*0.18))))])


        ws.append([""])
        ws.append([""])
        ws.append([""])
        ws.append([(''),(''),font_calibri_11_bold_upper(ws,'__________________'),(''),font_calibri_11_bold_upper(ws,'__________________')])
        ws.append([(''),(''),font_calibri_11_bold_upper(ws,'INSTRUMENTISTA'),(''),font_calibri_11_bold_upper(ws,'RECIBIDO POR')])
     
class SaleOrder(models.Model):
    _inherit = 'sale.order'

    def descargar_hoja_control(self):
	    
        import io
        output = io.BytesIO()

        workbook = Workbook(write_only=True)


        ws = workbook.create_sheet('Presupuesto_Pedido')

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 15
        ws.column_dimensions['R'].width = 15
        ws.column_dimensions['S'].width = 15
        ws.column_dimensions['T'].width = 15
        ws.column_dimensions['U'].width = 15
        ws.column_dimensions['V'].width = 15
        ws.column_dimensions['W'].width = 15



        ws.append([""])
        ws.merged_cells.ranges.append('D2:I2')
        tittle = font_calibri_16_bold(ws, "BIOCELLS GENOMICS SOCIEDAD ANONIMA CERRADA - BIOGENOMICS S.A.C.")
        ws.append([(''),(''),(''),tittle])
        ws.merged_cells.ranges.append('D3:I3')
        ruc = font_calibri_16_bold(ws, "RUC: 20607531600")
        ws.append([(''),(''),(''),ruc])
        ws.merged_cells.ranges.append('D4:I4')
        direccion = font_calibri_16_bold(ws, "AV. GUARDIA CIVIL NRO. 1321 DPTO. 1303 URB. VILLA VICTORIA - LIMA SURQUILLO")
        ws.append([(''),(''),(''),direccion])
        ws.merged_cells.ranges.append('D5:I5')
        telefono = font_calibri_16_bold(ws, "Telf: +51 1 6952357")
        ws.append([(''),(''),(''),telefono])
        ws.merged_cells.ranges.append('D6:I6')
        email = font_calibri_16_bold(ws, "Correo Electrónico: biocells.peru@biocells.pe ")
        ws.append([(''),(''),(''),email])
                                  

        ws.append([""])

        ws.merged_cells.ranges.append('B10:C10')
        fact = font_calibri_11_bold_upper(ws,'CÓDIGO DE CIRUGÍA')
        ws.append([(''),fact,(''),font_calibri_11(ws,': '+str(self.sugery_order)),(''),font_calibri_11_bold_upper(ws,'NOMBRE PACIENTE'),font_calibri_11(ws,': '+ str(self.name_patient))])
        ws.merged_cells.ranges.append('B11:C11')
        calle = font_calibri_11_bold_upper(ws,'FECHA DE CIRGUGÍA')
        ws.append([(''),calle,(''),font_calibri_11(ws,': '+str(self.sugery_date.strftime('%Y-%m-%d'))),(''),font_calibri_11_bold_upper(ws,'HISTORIA CLÍNICA'),font_calibri_11(ws,': '+ str(self.clinic_history))])
        ws.merged_cells.ranges.append('B12:C12')
        atencion = font_calibri_11_bold_upper(ws,'HORA DE CIRGUÍA')
        ws.append([(''),atencion,(''),font_calibri_11(ws,': '+str(self.sugery_date.strftime('%H:%M'))),(''),font_calibri_11_bold_upper(ws,'DNI/RUC'),font_calibri_11(ws,': '+ str(self.patient_vat))])
        ws.merged_cells.ranges.append('B13:C13')
        atencion = font_calibri_11_bold_upper(ws,'CIRUJANO')
        ws.append([(''),atencion,(''),font_calibri_11(ws,': '+str(self.name_doctor.name)),(''),font_calibri_11_bold_upper(ws,'PROCEDIMIENTO'),font_calibri_11(ws,': '+ str(self.procedure_clinic))])
        atencion = font_calibri_11_bold_upper(ws,'CENTRO MÉDICO')
        ws.append([(''),atencion,(''),font_calibri_11(ws,': '+str(self.medic_center if self.medic_center else '')),('')])

       
       
        ws.append([""])
        ws.append([(''),format_cel_calibri_11_bold_centered(ws,'CÓDIGO'),format_cel_calibri_11_bold_centered(ws,'DESCRIPCIÓN'),format_cel_calibri_11_bold_centered(ws,'LOTE'),format_cel_calibri_11_bold_centered(ws,'RS'),format_cel_calibri_11_bold_centered(ws,'CANTIDAD'),format_cel_calibri_11_bold_centered(ws,'P. UNIT.'),format_cel_calibri_11_bold_centered(ws,'P. TOTAL.')])
        subtotal=0
        for move in self.move_ids_return:
            for line in move.move_line_ids:
                if line.qty_consumida > 0:
                    ws.append([(''),format_cel_calibri_11_centered_1(ws,str(line.product_id.default_code) if line.product_id.default_code else " "),format_cel_calibri_11_number(ws, str(line.product_id.name)),format_cel_calibri_11_centered_1(ws,str(line.lot_id.name or '')),format_cel_calibri_11_centered_1(ws,str(line.x_studio_registro_sanitario or '')),format_cel_calibri_11_centered_1(ws,str(int(line.qty_consumida))),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(move.price_unit_sale))),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(line.qty_consumida * move.price_unit_sale)))])
                    subtotal = subtotal + (line.qty_consumida * move.price_unit_sale)
        ws.append([""])
        ws.append([(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"SUBTOTAL"),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(subtotal)))])
        ws.append([(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"IGV"),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(subtotal*0.18)))])
        ws.append([(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"TOTAL"),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(subtotal+(subtotal*0.18))))])


        ws.append([""])
        ws.append([""])
        ws.append([""])
        ws.append([(''),(''),font_calibri_11_bold_upper(ws,'_____________'),(''),font_calibri_11_bold_upper(ws,'_____________')])
        ws.append([(''),(''),font_calibri_11_bold_upper(ws,'INSTRUMENTISTA'),(''),font_calibri_11_bold_upper(ws,'RECIBIDO POR')])
        workbook.save(output)
        output.seek(0)

        return self.env['popup.it'].get_file('Prespuesto_Pedido.xlsx',base64.encodestring(output.read()))
