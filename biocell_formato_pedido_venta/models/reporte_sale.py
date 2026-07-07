# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
from io import BytesIO
from datetime import *
import base64
import subprocess
import sys

def install(package):
	subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
	import openpyxl
except:
	install('openpyxl==3.0.5')

import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl import Workbook
values = {}
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell


from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.colors import magenta, red , black , blue, gray, Color, HexColor
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Table
from reportlab.lib.units import  cm,mm
from reportlab.lib.utils import simpleSplit
import decimal

def font_calibri_16_bold(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=12, bold=True)
    return cell
def font_calibri_bold(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=16, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    return cell

def font_calibri_11_bold_upper(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=11, bold=True)
    return cell

def font_calibri_11(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=11, bold=False)
    return cell

def format_cel_calibri_11_bold_centered(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=11, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Configura el relleno de la celda en celeste claro
    fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    cell.fill = fill
    
    # Configura los bordes de la celda
    cell.border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'))
    
    return cell


def format_cel_calibri_11_centered(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=11, bold=False)
    
    # Configura los bordes de la celda
    cell.border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'))
    
    return cell

def format_cel_calibri_11_centered_1(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=11, bold=False)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Configura los bordes de la celda
    cell.border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'))
    
    return cell



def format_cel_calibri_11_number(ws, texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Calibri', size=11, bold=False)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = "0.00"  # Establece el formato de número a 2 decimales

    # Configura los bordes de la celda
    cell.border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000'))
    
    return cell



class biocell_formato_pedido_venta_biocell_formato_pedido_venta_xlsx(models.AbstractModel):
    _name = 'report.biocell_fpv.biocell_formato_pedido_venta_xlsx'
    _inherit = 'report.biocell_generador_excel.abstract'

    def generate_xlsx_report(self, workbook, data, so):

        ws = workbook.create_sheet('Presupuesto_Pedido')

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 15
        ws.column_dimensions['R'].width = 15
        ws.column_dimensions['S'].width = 15
        ws.column_dimensions['T'].width = 15
        ws.column_dimensions['U'].width = 15
        ws.column_dimensions['V'].width = 15
        ws.column_dimensions['W'].width = 15



        ws.append([""])
        ws.merged_cells.ranges.append('D2:I2')
        tittle = font_calibri_16_bold(ws, "BIOCELLS GENOMICS SOCIEDAD ANONIMA CERRADA - BIOGENOMICS S.A.C.")
        ws.append([(''),(''),(''),tittle])
        ws.merged_cells.ranges.append('D3:I3')
        ruc = font_calibri_16_bold(ws, "RUC: 20607531600")
        ws.append([(''),(''),(''),ruc])
        ws.merged_cells.ranges.append('D4:I4')
        direccion = font_calibri_16_bold(ws, "Av. Santa Maria 140 Int. 201, Lima, Perú -  CMS Grau")
        ws.append([(''),(''),(''),direccion])
        ws.merged_cells.ranges.append('D5:I5')
        telefono = font_calibri_16_bold(ws, "Telf: +51 1 6952357")
        ws.append([(''),(''),(''),telefono])
        ws.merged_cells.ranges.append('D6:I6')
        email = font_calibri_16_bold(ws, "Correo Electrónico: biocells.peru@biocells.pe ")
        ws.append([(''),(''),(''),email])
                                  

        ws.append([""])


        ws.merged_cells.ranges.append('B8:H8')
        venta = font_calibri_bold(ws, "PEDIDO - " + str(so.name))
        ws.append([(''),venta])

        ws.merged_cells.ranges.append('B9:D9')
        mes = so.date_order.strftime("%B")
        dia = so.date_order.strftime("%d")
        año = so.date_order.strftime("%Y")
        meses_español = {
            'January': 'Enero',
            'February': 'Febrero',
            'March': 'Marzo',
            'April': 'Abril',
            'May': 'Mayo',
            'June': 'Junio',
            'July': 'Julio',
            'August': 'Agosto',
            'September': 'Septiembre',
            'October': 'Octubre',
            'November': 'Noviembre',
            'December': 'Diciembre'
        }
        fecha = font_calibri_11_bold_upper(ws, "Lima, " + meses_español[mes] + " "+ dia + ", "+ año)
        ws.append([(''),fecha])

        ws.merged_cells.ranges.append('B10:C10')
        fact = font_calibri_11_bold_upper(ws,'Facturar A')
        ws.merged_cells.ranges.append('D10:H10')
        cli = font_calibri_11(ws,': '+ str(so.partner_id.name))
        ws.append([(''),fact,(''),cli])
        ws.merged_cells.ranges.append('B11:C11')
        calle = font_calibri_11_bold_upper(ws,'Calle')
        ws.append([(''),calle,(''),font_calibri_11(ws,': '+ str(so.partner_id.street))])
        ws.merged_cells.ranges.append('B12:C12')
        atencion = font_calibri_11_bold_upper(ws,'Atención')
        ws.append([(''),atencion,(''),font_calibri_11(ws,': '+str(so.user_id.name)),(''),font_calibri_11_bold_upper(ws,'Teléfono'),font_calibri_11(ws,': '+ str(so.partner_id.phone))])

        ws.append([(''),font_calibri_11_bold_upper(ws,'Distinguido/a.-:')])
        ws.merged_cells.ranges.append('B14:F14')
        contenido = font_calibri_11_bold_upper(ws, "De acuerdo a lo solicitado, enviamos nuestra cotización")
        ws.append([(''),contenido])
       
        ws.append([""])
        desc = False
        for line_d in so.order_line:
            if line_d.discount > 0:
                desc = True
            
        
        if desc:
            ws.append([(''),format_cel_calibri_11_bold_centered(ws,'Artículo'),format_cel_calibri_11_bold_centered(ws,'Código'),format_cel_calibri_11_bold_centered(ws,'Descripción'),format_cel_calibri_11_bold_centered(ws,'Lote'),format_cel_calibri_11_bold_centered(ws,'RS'),format_cel_calibri_11_bold_centered(ws,'Und.'),format_cel_calibri_11_bold_centered(ws,'Cant.'),format_cel_calibri_11_bold_centered(ws,'P. Unit.'),format_cel_calibri_11_bold_centered(ws,'Descuento'),format_cel_calibri_11_bold_centered(ws,'Total')])
        else:
            ws.append([(''),format_cel_calibri_11_bold_centered(ws,'Artículo'),format_cel_calibri_11_bold_centered(ws,'Código'),format_cel_calibri_11_bold_centered(ws,'Descripción'),format_cel_calibri_11_bold_centered(ws,'Lote'),format_cel_calibri_11_bold_centered(ws,'RS'),format_cel_calibri_11_bold_centered(ws,'Und.'),format_cel_calibri_11_bold_centered(ws,'Cant.'),format_cel_calibri_11_bold_centered(ws,'P. Unit.'),format_cel_calibri_11_bold_centered(ws,'Total')])

        
        counter = 0

        if desc:
            for line in so.order_line:
                if line.product_id:
                    counter += 1
                    lot_names = list({ml.lot_id.name for m in line.move_ids for ml in m.move_line_ids if ml.lot_id})
                    lot_display = ', '.join(lot_names)
                    rs_display = str(line.product_id.x_studio_registro_sanitario or '')
                    ws.append([(''),format_cel_calibri_11_centered_1(ws,str(counter)),format_cel_calibri_11_number(ws,str(line.product_id.default_code) if line.product_id.default_code else " "),format_cel_calibri_11_number(ws, str(line.product_id.name)),format_cel_calibri_11_centered_1(ws,lot_display),format_cel_calibri_11_centered_1(ws,rs_display),format_cel_calibri_11_centered_1(ws,str(line.product_uom.name)),format_cel_calibri_11_centered_1(ws,str(int(line.product_uom_qty))),format_cel_calibri_11_number(ws,str(so.currency_id.symbol)+' '+ str("{:.2f}".format(line.price_unit))),format_cel_calibri_11_number(ws,str("{:.2f}".format(line.discount if line.discount else 0.00))+ ' %'),format_cel_calibri_11_number(ws,str(so.currency_id.symbol)+ ' ' + str("{:.2f}".format(line.price_subtotal)))])
        else:
            for line in so.order_line:
                if line.product_id:
                    counter += 1
                    lot_names = list({ml.lot_id.name for m in line.move_ids for ml in m.move_line_ids if ml.lot_id})
                    lot_display = ', '.join(lot_names)
                    rs_display = str(line.product_id.x_studio_registro_sanitario or '')
                    ws.append([(''),format_cel_calibri_11_centered_1(ws,str(counter)),format_cel_calibri_11_number(ws,str(line.product_id.default_code) if line.product_id.default_code else " "),format_cel_calibri_11_number(ws, str(line.product_id.name)),format_cel_calibri_11_centered_1(ws,lot_display),format_cel_calibri_11_centered_1(ws,rs_display),format_cel_calibri_11_centered_1(ws,str(line.product_uom.name)),format_cel_calibri_11_centered_1(ws,str(int(line.product_uom_qty))),format_cel_calibri_11_number(ws,str(so.currency_id.symbol)+' ' + str("{:.2f}".format(line.price_unit))),format_cel_calibri_11_number(ws,str(so.currency_id.symbol)+' ' + str("{:.2f}".format(line.price_subtotal)))])
        
        if desc:
            ws.append([""])
            ws.append([(''),(''),(''),(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"Sub-Total"),format_cel_calibri_11_number(ws, str(so.currency_id.symbol) +' ' + str("{:.2f}".format(so.amount_untaxed)))])
            ws.append([(''),(''),(''),(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"IGV (18%)"),format_cel_calibri_11_number(ws,str(so.currency_id.symbol) +' ' + str("{:.2f}".format(so.amount_tax)))])
            ws.append([(''),(''),(''),(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"Total ("+ str(so.currency_id.symbol)+")"),format_cel_calibri_11_number(ws,str(so.currency_id.symbol) +' ' + str("{:.2f}".format(so.amount_total)))])

        else:
            ws.append([""])
            ws.append([(''),(''),(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"Sub-Total"),format_cel_calibri_11_number(ws, str(so.currency_id.symbol) +' ' + str("{:.2f}".format(so.amount_untaxed)))])
            ws.append([(''),(''),(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"IGV (18%)"),format_cel_calibri_11_number(ws,str(so.currency_id.symbol) +' ' + str("{:.2f}".format(so.amount_tax)))])
            ws.append([(''),(''),(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"Total ("+ str(so.currency_id.symbol)+")"),format_cel_calibri_11_number(ws,str(so.currency_id.symbol) +' ' + str("{:.2f}".format(so.amount_total)))])
        
           

        ws.append([""])

        ws.append([(''),font_calibri_11(ws,'Nombre del paciente : '+str(so.name_patient if so.name_patient else " "))])
        ws.append([(''),font_calibri_11(ws,'Nombre del doctor : '+str(so.name_doctor.name if so.name_doctor.name else " "))])
        ws.append([""])
        ws.append([(''),font_calibri_11_bold_upper(ws,'Condiciones')])
        ws.append([(''),font_calibri_11(ws,'Forma de pago : '+str(so.payment_term_id.name if so.payment_term_id else " "))])
        policy ={
             'direct':'Lo antes posible',
             'one':'Cuando todos los productos estén listos'
        }
        ws.append([(''),font_calibri_11(ws,'Tiempo de entrega: '+str(policy[so.picking_policy] if so.picking_policy else " "))])
        ws.append([(''),font_calibri_11(ws,'Validez de la oferta : '+str(so.validity_date.strftime('%d/%m/%Y') if so.validity_date else " "))])
        
        ws.append([(''),font_calibri_11(ws,'Sin otro particular y esperando su favorable respuesta, me despido.')])
        ws.append([(''),font_calibri_11(ws,'Atentamente,')])
        ws.append([""])
        ws.append([""])
        ws.append([(''),(''),(''),font_calibri_11_bold_upper(ws,'Agregar firma de usuario')])
        ws.append([""])
        ws.append([""])
        ws.append([(''),(''),(''),font_calibri_11_bold_upper(ws,'_____________________')])
        ws.append([(''),font_calibri_11_bold_upper(ws,'Pagar a:')])
        ws.append([(''),font_calibri_11_bold_upper(ws,'N° de cuenta corriente')])
        ws.append([(''),font_calibri_11(ws,'BANCO DE CREDITO DEL PERU S/ : 193-2647415-0-64')])
        ws.append([(''),font_calibri_11(ws,'BANCO DE CREDITO DEL PERU $ : 193-9294084-1-44')])
        ws.append([(''),font_calibri_11(ws,'BANCO DE LA NACION S/ : 00-008-097828')])





class SaleOrder(models.Model):
    _inherit = 'sale.order'

    def descargar_report_ventas(self):
	    
        import io
        output = io.BytesIO()

        workbook = Workbook(write_only=True)


        ws = workbook.create_sheet('Presupuesto_Pedido')

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 15
        ws.column_dimensions['R'].width = 15
        ws.column_dimensions['S'].width = 15
        ws.column_dimensions['T'].width = 15
        ws.column_dimensions['U'].width = 15
        ws.column_dimensions['V'].width = 15
        ws.column_dimensions['W'].width = 15



        ws.append([""])
        ws.merged_cells.ranges.append('D2:I2')
        tittle = font_calibri_16_bold(ws, "BIOCELLS GENOMICS SOCIEDAD ANONIMA CERRADA - BIOGENOMICS S.A.C.")
        ws.append([(''),(''),(''),tittle])
        ws.merged_cells.ranges.append('D3:I3')
        ruc = font_calibri_16_bold(ws, "RUC: 20607531600")
        ws.append([(''),(''),(''),ruc])
        ws.merged_cells.ranges.append('D4:I4')
        direccion = font_calibri_16_bold(ws, "Av. Santa Maria 140 Int. 201, Lima, Perú -  CMS Grau")
        ws.append([(''),(''),(''),direccion])
        ws.merged_cells.ranges.append('D5:I5')
        telefono = font_calibri_16_bold(ws, "Telf: +51 1 6952357")
        ws.append([(''),(''),(''),telefono])
        ws.merged_cells.ranges.append('D6:I6')
        email = font_calibri_16_bold(ws, "Correo Electrónico: biocells.peru@biocells.pe ")
        ws.append([(''),(''),(''),email])
                                  

        ws.append([""])


        ws.merged_cells.ranges.append('B8:H8')
        venta = font_calibri_bold(ws, "PEDIDO - " + str(self.name))
        ws.append([(''),venta])

        ws.merged_cells.ranges.append('B9:D9')
        mes = self.date_order.strftime("%B")
        dia = self.date_order.strftime("%d")
        año = self.date_order.strftime("%Y")
        meses_español = {
            'January': 'Enero',
            'February': 'Febrero',
            'March': 'Marzo',
            'April': 'Abril',
            'May': 'Mayo',
            'June': 'Junio',
            'July': 'Julio',
            'August': 'Agosto',
            'September': 'Septiembre',
            'October': 'Octubre',
            'November': 'Noviembre',
            'December': 'Diciembre'
        }
        fecha = font_calibri_11_bold_upper(ws, "Lima, " + meses_español[mes] + " "+ dia + ", "+ año)
        ws.append([(''),fecha])

        ws.merged_cells.ranges.append('B10:C10')
        fact = font_calibri_11_bold_upper(ws,'Facturar A')
        ws.merged_cells.ranges.append('D10:H10')
        cli = font_calibri_11(ws,': '+ str(self.partner_id.name))
        ws.append([(''),fact,(''),cli])
        ws.merged_cells.ranges.append('B11:C11')
        calle = font_calibri_11_bold_upper(ws,'Calle')
        ws.append([(''),calle,(''),font_calibri_11(ws,': '+ str(self.partner_id.street))])
        ws.merged_cells.ranges.append('B12:C12')
        atencion = font_calibri_11_bold_upper(ws,'Atención')
        ws.append([(''),atencion,(''),font_calibri_11(ws,': '+str(self.user_id.name)),(''),font_calibri_11_bold_upper(ws,'Teléfono'),font_calibri_11(ws,': '+ str(self.partner_id.phone))])

        ws.append([(''),font_calibri_11_bold_upper(ws,'Distinguido/a.-:')])
        ws.merged_cells.ranges.append('B14:F14')
        contenido = font_calibri_11_bold_upper(ws, "De acuerdo a lo solicitado, enviamos nuestra cotización")
        ws.append([(''),contenido])
       
        ws.append([""])
        ws.append([(''),format_cel_calibri_11_bold_centered(ws,'Artículo'),format_cel_calibri_11_bold_centered(ws,'Código'),format_cel_calibri_11_bold_centered(ws,'Descripción'),format_cel_calibri_11_bold_centered(ws,'Lote'),format_cel_calibri_11_bold_centered(ws,'RS'),format_cel_calibri_11_bold_centered(ws,'Und.'),format_cel_calibri_11_bold_centered(ws,'Cant.'),format_cel_calibri_11_bold_centered(ws,'P. Unit.'),format_cel_calibri_11_bold_centered(ws,'Total')])
        counter = 0
        for line in self.order_line_origin:
            if line.product_id:
                counter += 1
                lot_names = list({ml.lot_id.name for m in line.order_line_id.move_ids for ml in m.move_line_ids if ml.lot_id})
                lot_display = ', '.join(lot_names)
                rs_display = str(line.x_registro_sanitario or '')
                ws.append([(''),format_cel_calibri_11_centered_1(ws,str(counter)),format_cel_calibri_11_number(ws,str(line.product_id.default_code) if line.product_id.default_code else " "),format_cel_calibri_11_number(ws, str(line.product_id.name)),format_cel_calibri_11_centered_1(ws,lot_display),format_cel_calibri_11_centered_1(ws,rs_display),format_cel_calibri_11_centered_1(ws,str(line.product_uom.name)),format_cel_calibri_11_centered_1(ws,str(int(line.product_uom_qty))),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(line.price_unit))),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(line.price_subtotal)))])
        
        ws.append([""])
        ws.append([(''),(''),(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"Sub-Total"),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(self.amount_untaxed_origin)))])
        ws.append([(''),(''),(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"IGV (18%)"),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(self.amount_tax_origin)))])
        ws.append([(''),(''),(''),(''),(''),(''),(''),(''),format_cel_calibri_11_bold_centered(ws,"Total (S/.)"),format_cel_calibri_11_number(ws,"S/ " + str("{:.2f}".format(self.amount_total_origin)))])


        ws.append([""])

        ws.append([(''),font_calibri_11(ws,'Nombre del paciente : '+str(self.name_patient if self.name_patient else " "))])
        ws.append([(''),font_calibri_11(ws,'Nombre del doctor : '+str(self.name_doctor.name if self.name_doctor.name else " "))])
        ws.append([""])
        ws.append([(''),font_calibri_11_bold_upper(ws,'Condiciones')])
        ws.append([(''),font_calibri_11(ws,'Forma de pago : '+str(self.payment_term_id.name if self.payment_term_id else " "))])
        policy ={
             'direct':'Lo antes posible',
             'one':'Cuando todos los productos estén listos'
        }
        ws.append([(''),font_calibri_11(ws,'Tiempo de entrega: '+str(policy[self.picking_policy] if self.picking_policy else " "))])
        ws.append([(''),font_calibri_11(ws,'Validez de la oferta : '+str(self.validity_date.strftime('%d/%m/%Y') if self.validity_date else " "))])
        
        ws.append([(''),font_calibri_11(ws,'Sin otro particular y esperando su favorable respuesta, me despido.')])
        ws.append([(''),font_calibri_11(ws,'Atentamente,')])
        ws.append([""])
        ws.append([""])
        ws.append([(''),(''),(''),font_calibri_11_bold_upper(ws,'Agregar firma de usuario')])
        ws.append([""])
        ws.append([""])
        ws.append([(''),(''),(''),font_calibri_11_bold_upper(ws,'_____________________')])
        ws.append([(''),font_calibri_11_bold_upper(ws,'Pagar a:')])
        ws.append([(''),font_calibri_11_bold_upper(ws,'N° de cuenta corriente')])
        ws.append([(''),font_calibri_11(ws,'BANCO DE CREDITO DEL PERU S/ : 193-2647415-0-64')])
        ws.append([(''),font_calibri_11(ws,'BANCO DE CREDITO DEL PERU $ : 193-9294084-1-44')])
        ws.append([(''),font_calibri_11(ws,'BANCO DE LA NACION S/ : 00-008-097828')])



        workbook.save(output)
        output.seek(0)

        return self.env['popup.it'].get_file('Prespuesto_Pedido.xlsx',base64.encodestring(output.read()))
